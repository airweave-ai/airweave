"""Excel bongo implementation.

Creates, updates, and deletes test entities via the real Microsoft Graph API.
"""

import io
import uuid
from typing import Any, Dict, List, Optional

import httpx
from monke.bongos._microsoft_graph_base import GRAPH, MicrosoftGraphBongo
from monke.generation.excel import generate_workbook_content

# Try to import openpyxl for Excel file creation
try:
    from openpyxl import Workbook, load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelBongo(MicrosoftGraphBongo):
    """Bongo for Excel that creates test entities for E2E testing.

    Key responsibilities:
    - Create test Excel workbooks in OneDrive
    - Add worksheets with data containing verification tokens
    - Update worksheets to test incremental sync
    - Delete workbooks to test deletion detection
    - Clean up all test data
    """

    connector_type = "excel"

    def __init__(self, credentials: Dict[str, Any], **kwargs):
        super().__init__(credentials, **kwargs)

        # Track created resources for cleanup
        self._test_workbook_id: Optional[str] = None
        self._test_workbook_name: Optional[str] = None
        self._worksheets: List[Dict[str, Any]] = []

        if not HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required for Excel bongo. Install with: pip install openpyxl"
            )

    async def create_entities(self) -> List[Dict[str, Any]]:
        """Create test Excel workbook with worksheets in OneDrive."""
        self.logger.info(
            f"🥁 Creating Excel test workbook with {self.entity_count} worksheets"
        )
        out: List[Dict[str, Any]] = []

        # Generate tokens for each worksheet
        tokens = [uuid.uuid4().hex[:8] for _ in range(self.entity_count)]

        # Generate workbook content
        test_name = f"TestData_{uuid.uuid4().hex[:8]}"
        filename, worksheet_data = await generate_workbook_content(
            self.openai_model, tokens, test_name
        )
        self._test_workbook_name = filename

        self.logger.info(f"📊 Generated {len(worksheet_data)} worksheets")

        # Create Excel workbook file
        safe_filename = self._sanitize_filename(filename)
        wb_bytes = self._create_excel_file(worksheet_data, tokens)

        # Upload to OneDrive
        async with httpx.AsyncClient(base_url=GRAPH, timeout=60) as client:
            await self._pace()
            self.logger.info(f"📤 Uploading Excel workbook: {safe_filename}")

            upload_url = f"/me/drive/root:/{safe_filename}:/content"
            r = await client.put(
                upload_url,
                headers={
                    "Authorization": f"Bearer {self.access_token}",
                    "Content-Type": (
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    ),
                },
                content=wb_bytes,
            )

            if r.status_code not in (200, 201):
                self.logger.error(f"Upload failed {r.status_code}: {r.text}")
                r.raise_for_status()

            wb_file = r.json()
            self._test_workbook_id = wb_file["id"]

            self.logger.info(f"✅ Uploaded workbook: {self._test_workbook_id} - {safe_filename}")

        # Build entity list for verification
        for i, (sheet_data, token) in enumerate(zip(worksheet_data, tokens)):
            ent = {
                "type": "worksheet",
                "workbook_id": self._test_workbook_id,
                "workbook_name": safe_filename,
                "sheet_name": sheet_data.name,  # Use attribute access for Pydantic model
                "token": token,
                "expected_content": token,
            }
            out.append(ent)
            self._worksheets.append(ent)

        self.created_entities.append(
            {"id": self._test_workbook_id, "name": safe_filename}
        )

        self.logger.info(
            f"✅ Created workbook with {len(self._worksheets)} worksheets containing tokens"
        )
        return out

    def _create_excel_file(
        self, worksheet_data: List[Dict[str, Any]], tokens: List[str]
    ) -> bytes:
        """Create an Excel file with the given worksheet data and tokens.

        Args:
            worksheet_data: List of worksheet content dicts
            tokens: List of verification tokens

        Returns:
            Bytes of the Excel workbook
        """
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create worksheets
        for i, (sheet_data, token) in enumerate(zip(worksheet_data, tokens)):
            ws = wb.create_sheet(title=sheet_data.name)  # Use attribute access for Pydantic model

            # Write headers
            headers = sheet_data.headers  # Use attribute access for Pydantic model
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # Write data rows with token
            for row_idx, row_data in enumerate(sheet_data.rows, start=2):  # Use attribute access for Pydantic model
                for col_idx, value in enumerate(row_data, start=1):
                    # Include token in first data row
                    if row_idx == 2 and col_idx == 1:
                        value = f"{value} ({token})"
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    async def update_entities(self) -> List[Dict[str, Any]]:
        """Update Excel workbook by adding data to worksheets."""
        if not self._test_workbook_id:
            return []

        self.logger.info("🥁 Updating worksheets in Excel workbook")
        updated = []

        async with httpx.AsyncClient(base_url=GRAPH, timeout=60) as client:
            # Download current workbook
            await self._pace()
            download_url = f"/me/drive/items/{self._test_workbook_id}/content"
            r = await client.get(download_url, headers=self._hdrs())

            if r.status_code != 200:
                self.logger.warning(
                    f"Failed to download workbook: {r.status_code} - {r.text[:200]}"
                )
                return []

            # Load existing workbook from downloaded content
            wb = load_workbook(io.BytesIO(r.content))

            # Add a new row to each worksheet with updated data
            for ent in self._worksheets[: min(2, len(self._worksheets))]:
                sheet_name = ent['sheet_name']
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # Append a new row with updated marker
                    next_row = ws.max_row + 1
                    ws.cell(row=next_row, column=1, value=f"Updated: {ent['token']}")
                    ws.cell(row=next_row, column=2, value="Modified by Monke test")
                    
                    updated.append({**ent, "updated": True})
                    self.logger.info(
                        f"📝 Added update row to worksheet '{sheet_name}' with token: {ent['token']}"
                    )

            # Save updated workbook
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Upload updated workbook back to OneDrive
            await self._pace()
            upload_url = f"/me/drive/items/{self._test_workbook_id}/content"
            r = await client.put(
                upload_url,
                headers={
                    "Authorization": f"Bearer {self.access_token}",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
                content=buffer.getvalue(),
            )

            if r.status_code in (200, 201):
                self.logger.info("✅ Successfully uploaded updated workbook")
            else:
                self.logger.warning(f"Failed to upload updated workbook: {r.status_code}")

        return updated

    async def delete_entities(self) -> List[str]:
        """Delete all test entities (workbook)."""
        if not self._test_workbook_id:
            return []

        self.logger.info("🥁 Deleting Excel test workbook")
        deleted: List[str] = []

        async with httpx.AsyncClient(base_url=GRAPH, timeout=30) as client:
            success = await self._delete_with_retry(
                client, self._test_workbook_id, self._test_workbook_name or "TestWorkbook"
            )

            if success:
                deleted.append(self._test_workbook_id)
                self.logger.info(f"✅ Deleted workbook: {self._test_workbook_name}")

                # Clear tracking
                self._test_workbook_id = None
                self._test_workbook_name = None
                self._worksheets = []

        return deleted

    async def delete_specific_entities(self, entities: List[Dict[str, Any]]) -> List[str]:
        """Delete specific worksheets (not implemented - deletes entire workbook)."""
        # Excel bongo deletes the entire workbook, not individual worksheets
        return await self.delete_entities()

    async def cleanup(self):
        """Comprehensive cleanup of all test resources."""
        self.logger.info("🧹 Starting comprehensive Excel cleanup")

        cleanup_stats = {
            "workbooks_deleted": 0,
            "files_deleted": 0,
            "errors": 0,
        }

        try:
            async with httpx.AsyncClient(base_url=GRAPH, timeout=30) as client:
                # Delete tracked test workbook
                if self._test_workbook_id:
                    self.logger.info("🗑️  Deleting tracked workbook")
                    deleted = await self.delete_entities()
                    cleanup_stats["workbooks_deleted"] += len(deleted)

                # Search for and cleanup any orphaned test workbooks
                await self._cleanup_orphaned_files(client, cleanup_stats, "Monke_", [".xlsx"])

            self.logger.info(
                f"🧹 Cleanup completed: {cleanup_stats['workbooks_deleted']} "
                f"workbooks deleted, {cleanup_stats['files_deleted']} orphaned files deleted, "
                f"{cleanup_stats['errors']} errors"
            )
        except Exception as e:
            self.logger.error(f"❌ Error during comprehensive cleanup: {e}")
